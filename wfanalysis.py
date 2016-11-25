import openpyxl
import numpy as np
from sklearn import cluster, datasets
import matplotlib.pyplot as plt
from itertools import cycle
import scipy.stats as st
import math
ncustomers=10000
filename='wfmindsumo.xlsx'
################################################################################################
def getdata():
	wb = openpyxl.load_workbook(filename,data_only=True)
	datasheet = wb.get_sheet_by_name('Data')
	return datasheet
################################
def clusterbytotalgrowth(datasheet):
	totalABaccountgrowth="";
	yeartotalgrowth="";
	nrows=datasheet.max_row
	ncolumns=datasheet.max_column
	for row in datasheet.iter_rows(min_row=1, max_col=ncolumns, max_row=1):
		for cell in row:
			if(cell.value=='TotalABaccountgrowth'):
				abletter= cell.column;
				totalABaccountgrowth=datasheet[abletter];
			if(cell.value=='yeartotalgrowth'):
				letter= cell.column;
				yeartotalgrowth=datasheet[letter];
	ABaccountgrowthdata=[]
	for cell in totalABaccountgrowth:
		if (cell.value !='#N/A'):
			ABaccountgrowthdata.append(cell.value)
	yeartotalgrowthdata=[]
	for cell in yeartotalgrowth:
		if (cell.value !='#N/A'):
			yeartotalgrowthdata.append(cell.value)
	kmeans=cluster.KMeans(n_clusters=5)
	ABaccountgrowthdata.pop(0)
	yeartotalgrowthdata.pop(0)
	clusterdata=np.column_stack((ABaccountgrowthdata,yeartotalgrowthdata))
	kmeans.fit(clusterdata)
	return [clusterdata,kmeans]
################################################################################
def plotclusters(kmeans,clusterdata):
	labels=kmeans.labels_
	cluster_centers=kmeans.cluster_centers_
	labels_unique=np.unique(labels)
	n_clusters_=len(labels_unique)
	print("number of estimated clusters : %d" % n_clusters_)
	plt.figure(1)
	plt.clf()
	#Warning: graphics only cogent when colors pattern number is the same as the amount of 		clusters
	#ex:'bgrcm' is 5 letters, and cluster number is 5	
	colors = cycle('bgrcmbgrcmbgrcm')
	for k, col in zip(range(n_clusters_), colors):
	    my_members = labels == k
	    cluster_center = cluster_centers[k]
	    plt.plot(clusterdata[my_members, 0], clusterdata[my_members, 1], col + '.')
	    plt.plot(cluster_center[0], cluster_center[1], 'o', markerfacecolor=col,
	             markeredgecolor='k', markersize=14)
	plt.title('Estimated number of clusters: %d' % n_clusters_)
	plt.show()
#########################################################################################
def gettargetlist(kmeans,abthreshhold,balancethreshold):
	clustercount=0
	interestingclusterlist=[]
	labels=kmeans.labels_
	cluster_centers=kmeans.cluster_centers_
	for i in cluster_centers:
		if(i[0]>abthreshhold and i[1]>balancethreshold):
			#
			interestingclusterlist.append(clustercount)
		clustercount+=1
	targetlists= {i: np.where(kmeans.labels_ == i)[0] for i in range(kmeans.n_clusters)}
	indextrack=0
	targetlist=[]
	for i in labels:
		if(i==interestingclusterlist[0]):
			targetlist.append(indextrack)
		indextrack+=1
	return targetlist
###############################################################################################################333
def categoryclusterpval(targetlist,numparameters,testvariable,ncustomers,datasheet):
	print("Testing: " + testvariable)
	popdem1sum=[0]*numparameters
	subdem1sum=[0]*numparameters
	variableletter=""
	for cell in datasheet[1]:
		if(cell.value==testvariable):
			variableletter=cell.column
			break
	for i in range(1,ncustomers+1):
		aicell=datasheet[variableletter+str(i*12 +1)]
		popdem1sum[aicell.value-1]+=1
	for i in targetlist:
		cnum=i+1
		aicell=datasheet[variableletter+str(cnum*12+1)]
		subdem1sum[aicell.value-1]+=1
	popdem1propavg=[x/ float(ncustomers) for x in popdem1sum]
	subdem1propavg=[x/ float(len(targetlist)) for x in subdem1sum]
	print(popdem1propavg)
	print(subdem1propavg)
	tsdem1=[0]*numparameters
	dem1pval=[0]*numparameters
	for i in range(0,numparameters):
		p1ai=popdem1propavg[i]
		p2ai=subdem1propavg[i]
		if(p1ai==0 and p2ai ==0):
			tsdem1=0
			dem1pval=1
			break
		tsai=(p2ai-p1ai)/math.sqrt((p1ai*(1-p1ai)/ncustomers)+(p2ai*(1-p2ai)/len(targetlist)))
		tsdem1[i]=tsai
		dem1pval[i]= st.norm.sf(abs(tsdem1[i]))*2
	return [tsdem1,dem1pval]
################################################################################################################
def numericclusterpval(targetlist,testvariable,ncustomers,datasheet):
	print("testing: " + testvariable)
	popdata=[]
	subdata=[]
	variableletter=""		
	for cell in datasheet[1]:
		if(cell.value==testvariable):
			variableletter=cell.column
			break
	for i in range(1,ncustomers+1):
		popcell=datasheet[variableletter+str(i*12 +1)]
		popdata.append(popcell.value)
	for i in targetlist:
		cnum=i+1
		subcell=datasheet[variableletter+str(cnum*12+1)]
		subdata.append(subcell.value)
	popavg=np.mean(popdata)
	subavg=np.mean(subdata)
	popstd=np.std(popdata)
	substd=np.std(subdata)
	popvar=math.pow(popstd,2)
	subvar=math.pow(substd,2)
	print(subavg)
	print(popavg)
	ts= (subavg-popavg)/math.sqrt((popvar/ncustomers)+(subvar/len(targetlist)))
	pval=st.norm.sf(abs(ts))*2
	return [ts,pval]	
def findsignificance(ts,pval,threshold,startsat):
	if(type(ts) is not list):
		print(pval)
		if(pval<threshold):
			print(str("variable is significant:"))
			if (ts>0):
				print("Has positive effect")
				return
			else:
				print("Has negative effect")
				return
		print('no significance')
	else:
		for i in range(0,len(pval)):
			if(pval[i]<threshold):
				print(str(i+startsat) + " is significant:")
				if (ts[i]>0):
					print("	Has positive effect")
					return
				else:
					print("	Has negative effect")
					return
		print('no significance')
####################################################################################################

#Test Code:
datasheet=getdata()
clusterdata=clusterbytotalgrowth(datasheet)[0]
kmeans=clusterbytotalgrowth(datasheet)[1]
#uncomment below to see clusterplot
#plotclusters(kmeans,clusterdata)   
targetlist=gettargetlist(kmeans,0,1)
pvalaistatistics=categoryclusterpval(targetlist,5,"cust_demographics_ai",ncustomers,datasheet)
findsignificance(pvalaistatistics[0],pvalaistatistics[1],.05,1)
pvalstatistics=categoryclusterpval(targetlist,5,"cust_demographics_aii",ncustomers,datasheet)
findsignificance(pvalstatistics[0],pvalstatistics[1],.05,1)
variabletests=['aitotaloutreach','aiitotaloutreach','aiiitotaloutreach','aivtotaloutreach','avtotaloutreach','avitotaloutreach','aviitotaloutreach','aviiitotaloutreach','channelitotal','channeliitotal','channelivtotal']
for i in variabletests:
	testing=numericclusterpval(targetlist,i,ncustomers,datasheet)
	findsignificance(testing[0],testing[1],.05,0)
exposuretests=	['aifirstexposure','aiifirstexposure','aiiifirstexposure','aivfirstexposure','avfirstexposure','avifirstexposure','aviifirstexposure','aviiifirstexposure','channelifirstexposure','channeliifirstexposure','channeliiifirstexposure','channelivfirstexposure']
for i in exposuretests:
	testing=categoryclusterpval(targetlist,5,i,ncustomers,datasheet)
	findsignificance(testing[0],testing[1],.05,0)